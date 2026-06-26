# -*- coding: utf-8 -*-
"""生成《江苏2026本科招生专业目录（全国高校）》Excel。

数据：
  - data/jiangsu_plan_2026.json    （逐校×专业 招生计划，fetch_jiangsu_plan_2026.py 产出）
  - data/undergrad_majors_2026.json （全国本科专业总名录，参考用）

工作表：
  1. 江苏2026本科招生计划(明细)  —— 主表：院校×专业组×专业×计划数×选科×学费
  2. 院校计划汇总               —— 每所院校的专业数与计划合计
  3. 全国本科专业目录(2026)      —— 教育部本科专业总名录（883，参考）
  4. 说明与数据来源             —— 出处、覆盖范围与口径

用法：python crawler/build_jiangsu_excel.py [输出.xlsx]
"""
from __future__ import annotations
import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
PLAN_FILE = DATA / "jiangsu_plan_2026.json"
MAJORS_FILE = DATA / "undergrad_majors_2026.json"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="1F4E78")
SUB_FONT = Font(name="微软雅黑", bold=True, size=12)
CELL_FONT = Font(name="微软雅黑", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def write_detail(wb, plan):
    rows = plan["rows"]
    ws = wb.active
    ws.title = "江苏2026本科招生计划(明细)"
    headers = ["序号", "院校名称", "院校所在省", "院校国标代码", "批次", "科类",
               "专业组", "选科要求", "专业名称", "计划数", "学制", "学费(元/年)", "备注"]
    ws.append(headers)
    style_header(ws, len(headers))
    # 排序：科类、院校、专业组、计划数降序
    rows = sorted(rows, key=lambda r: (r.get("subject_type", ""), r.get("school_name", ""),
                                       str(r.get("group_name", "")),
                                       -(int(r["num"]) if str(r.get("num")).isdigit() else 0)))
    for i, r in enumerate(rows, start=1):
        ws.append([i, r["school_name"], r["school_province"], r["school_code"],
                   r["batch"], r["subject_type"], r["group_name"], r["select_req"],
                   r["special_name"],
                   int(r["num"]) if str(r["num"]).isdigit() else r["num"],
                   r["length"], r["tuition"], r["remark"]])
    # 数据区字体（大表，仅设字体+对齐，省略逐格边框以保证性能）
    left_cols = {2, 8, 9, 13}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = CELL_FONT
            cell.alignment = LEFT if cell.column in left_cols else CENTER
    widths = [6, 26, 10, 12, 12, 8, 8, 20, 30, 8, 8, 11, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    return len(rows)


def write_school_summary(wb, plan):
    ws = wb.create_sheet("院校计划汇总")
    agg = defaultdict(lambda: {"prov": "", "n_major": 0, "n_plan": 0})
    for r in plan["rows"]:
        k = r["school_name"]
        agg[k]["prov"] = r["school_province"]
        agg[k]["n_major"] += 1
        if str(r["num"]).isdigit():
            agg[k]["n_plan"] += int(r["num"])
    headers = ["序号", "院校名称", "所在省", "专业(组内)数", "计划合计(人)"]
    ws.append(headers)
    style_header(ws, len(headers))
    data = sorted(agg.items(), key=lambda kv: -kv[1]["n_plan"])
    for i, (name, v) in enumerate(data, start=1):
        ws.append([i, name, v["prov"], v["n_major"], v["n_plan"]])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = LEFT if cell.column == 2 else CENTER
    for i, w in enumerate([6, 30, 10, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def write_national_catalog(wb):
    if not MAJORS_FILE.exists():
        return
    obj = json.loads(MAJORS_FILE.read_text(encoding="utf-8"))
    ws = wb.create_sheet("全国本科专业目录(2026)")
    headers = ["序号", "学科门类", "专业类", "专业代码", "专业名称", "专业类型"]
    ws.append(headers)
    style_header(ws, len(headers))
    for m in obj["majors"]:
        ws.append([m["seq"], m["category"], m["major_class"], m["code"], m["name"], m["type"]])
    left_cols = {3, 5}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = LEFT if cell.column in left_cols else CENTER
    for i, w in enumerate([6, 12, 22, 12, 30, 24], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def write_readme(wb, plan, n_detail):
    ws = wb.create_sheet("说明与数据来源")
    ws.column_dimensions["A"].width = 118
    total_plan = sum(int(r["num"]) for r in plan["rows"] if str(r["num"]).isdigit())
    lines = [
        ("《江苏2026本科招生专业目录（全国高校）》数据说明", TITLE_FONT),
        ("", None),
        ("一、主表：江苏2026本科招生计划(明细)", SUB_FONT),
        (f"· 收录在江苏招生的全国高校 {plan['school_count_with_plan']} 所、专业明细 {n_detail} 行，"
         f"计划合计约 {total_plan:,} 人。", CELL_FONT),
        ("· 字段：院校名称 / 院校所在省 / 院校国标代码 / 批次 / 科类(首选历史·物理) / 专业组 / "
         "选科要求 / 专业名称 / 计划数 / 学制 / 学费 / 备注。", CELL_FONT),
        ("· 覆盖范围：普通类本科（历史等科目类 + 物理等科目类）为主，含少量职业本科。"
         "未包含：艺术类、体育类、强基计划、综合评价、高水平运动队、保送生等特殊类型"
         "（其逐校明细数据源未公开）。", Font(name="微软雅黑", size=10, color="C00000")),
        ("", None),
        ("二、数据来源与口径", SUB_FONT),
        ("· 来源：掌上高考（中国教育在线）公开数据接口 static-data.gaokao.cn（schoolspecialplan，2026/江苏）。", CELL_FONT),
        ("· 院校国标代码为教育部 5 位代码（如南京大学 10284），并非江苏志愿填报代码；"
         "江苏志愿填报院校/专业组/专业代码请以考试院系统为准。", CELL_FONT),
        ("· 权威口径：以江苏省教育考试院《江苏招生考试2026招生计划专刊》及官方志愿填报系统"
         "（gk.jseea.cn）为准。本表为第三方聚合整理，供检索与初步参考，正式填报务必逐条核对官方数据。", Font(name="微软雅黑", size=10, color="C00000")),
        ("· 官方汇总参考：1620 所高校在江苏安排计划 387657 人，其中本科 268542 人"
         "（江苏省教育考试院 2026-06-23 公布）。", CELL_FONT),
        ("", None),
        ("三、参考表：全国本科专业目录(2026)", SUB_FONT),
        ("· 教育部《普通高等学校本科专业目录（2025年）》(教高函〔2025〕3号) 全国本科专业总名录，883 个专业，"
         "2026 年招生执行。供了解专业全集与代码，非某校/某省的实际招生清单。", CELL_FONT),
    ]
    r = 1
    for text, font in lines:
        ws.cell(row=r, column=1, value=text)
        if font:
            ws.cell(row=r, column=1).font = font
        ws.cell(row=r, column=1).alignment = LEFT
        r += 1


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "江苏2026本科招生专业目录_全国高校.xlsx"
    plan = json.loads(PLAN_FILE.read_text(encoding="utf-8"))
    wb = Workbook()
    n_detail = write_detail(wb, plan)
    write_school_summary(wb, plan)
    write_national_catalog(wb)
    write_readme(wb, plan, n_detail)
    wb.save(out)
    print(f"已生成：{out}")
    print(f"明细 {n_detail} 行 / 院校 {plan['school_count_with_plan']} 所 / 工作表 {len(wb.sheetnames)} 个")
    print("工作表：", " / ".join(wb.sheetnames))


if __name__ == "__main__":
    main()
