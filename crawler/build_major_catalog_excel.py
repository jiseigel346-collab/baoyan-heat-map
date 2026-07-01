# -*- coding: utf-8 -*-
"""根据 data/undergrad_majors_2026.json 生成《2026本科招生专业目录》Excel。

包含 4 个工作表：
  1. 全国本科专业目录(2026)  —— 教育部本科专业目录（883 个专业，可核验）
  2. 学科门类统计           —— 按 12 个学科门类汇总
  3. 江苏2026招生计划(汇总)  —— 江苏省教育考试院官方公布的汇总数据
  4. 说明与数据来源         —— 出处、口径与局限说明

用法：python crawler/build_major_catalog_excel.py [输出路径.xlsx]
"""
from __future__ import annotations
import json
import sys
from collections import Counter, OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
MAJORS_FILE = DATA / "undergrad_majors_2026.json"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="1F4E78")
CELL_FONT = Font(name="微软雅黑", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 学科门类标准排序
CATEGORY_ORDER = ["哲学", "经济学", "法学", "教育学", "文学", "历史学", "理学",
                  "工学", "农学", "医学", "管理学", "艺术学", "交叉学科"]


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def write_catalog_sheet(wb, majors):
    ws = wb.active
    ws.title = "全国本科专业目录(2026)"
    headers = ["序号", "学科门类", "专业类", "专业代码", "专业名称", "专业类型"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    aligns = [CENTER, CENTER, LEFT, CENTER, LEFT, CENTER]
    for m in majors:
        ws.append([m["seq"], m["category"], m["major_class"], m["code"],
                   m["name"], m["type"]])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = aligns[c - 1]
    widths = [6, 12, 22, 12, 30, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def write_category_sheet(wb, majors):
    ws = wb.create_sheet("学科门类统计")
    counts = Counter(m["category"] for m in majors)
    type_counts = Counter(m["type"] for m in majors)
    ws.append(["学科门类", "专业数"])
    style_header(ws, 1, 2)
    total = 0
    for cat in CATEGORY_ORDER:
        if cat in counts:
            ws.append([cat, counts[cat]])
            total += counts[cat]
    ws.append(["合计", total])
    for r in range(2, ws.max_row + 1):
        for c in range(1, 3):
            cell = ws.cell(row=r, column=c)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = CENTER
    last = ws.max_row
    ws.cell(row=last, column=1).font = Font(name="微软雅黑", bold=True, size=10)
    ws.cell(row=last, column=2).font = Font(name="微软雅黑", bold=True, size=10)

    # 专业类型统计（错开两列展示）
    start = 1
    ws.cell(row=start, column=4, value="专业类型").fill = HEADER_FILL
    ws.cell(row=start, column=5, value="专业数").fill = HEADER_FILL
    for c in (4, 5):
        cell = ws.cell(row=start, column=c)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    r = start + 1
    for t, n in sorted(type_counts.items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=4, value=t)
        ws.cell(row=r, column=5, value=n)
        for c in (4, 5):
            cell = ws.cell(row=r, column=c)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = CENTER
        r += 1
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 10


def write_jiangsu_sheet(wb):
    ws = wb.create_sheet("江苏2026招生计划(汇总)")
    src = ("江苏省教育考试院 2026-06-23 公布；详见《江苏招生考试2026招生计划专刊》。"
           "来源：https://gaokao.chsi.com.cn/gkxx/zc/ss/202606/20260623/2293843281.html")
    ws.append(["江苏省2026年普通高校招生计划（官方汇总）"])
    ws.merge_cells("A1:C1")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = LEFT
    ws.append(["指标", "数值(人/所)", "口径说明"])
    style_header(ws, 2, 3)

    rows = [
        ("高考报名人数", "约521000", "全省2026年高考报名总数（52.1万）"),
        ("安排招生高校数", "1620", "在江苏安排统招计划的高等学校数（所）"),
        ("统招计划总数", "387657", "全国高校在江苏统考招生计划主要部分"),
        ("　本科计划", "268542", "含高校专项；不含强基/综评/高水平运动队/保送"),
        ("　专科计划", "119115", "统招专科"),
        ("高职提前招生专科计划", "127414", "前期已单独下达，未计入上面387657"),
        ("普通类合计", "351096", "历史科目类86131；物理科目类264965"),
        ("　普通类·历史·提前录取本科", "1441", ""),
        ("　普通类·历史·本科院校", "41881", ""),
        ("　普通类·历史·高职(专科)", "42809", ""),
        ("　普通类·物理·提前录取本科", "6182", ""),
        ("　普通类·物理·本科院校", "191576", ""),
        ("　普通类·物理·高职(专科)", "67207", ""),
        ("体育类合计", "3410", "历史科目类2107；物理科目类1303"),
        ("　体育类·历史·提前本科", "1369", ""),
        ("　体育类·历史·高职(专科)", "738", ""),
        ("　体育类·物理·提前本科", "883", ""),
        ("　体育类·物理·高职(专科)", "420", ""),
        ("艺术类合计", "33151", "历史科目类30054；物理科目类3097"),
        ("　艺术类·历史·提前本科", "23033", ""),
        ("　艺术类·历史·高职(专科)", "7021", ""),
        ("　艺术类·物理·提前本科", "2177", ""),
        ("　艺术类·物理·高职(专科)", "920", ""),
    ]
    for item, val, note in rows:
        ws.append([item, val, note])
    for r in range(3, ws.max_row + 1):
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = LEFT if c != 2 else CENTER
    ws.append([])
    note_row = ws.max_row + 1
    ws.cell(row=note_row, column=1, value=src)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)
    ws.cell(row=note_row, column=1).font = Font(name="微软雅黑", size=9, color="808080")
    ws.cell(row=note_row, column=1).alignment = LEFT
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 46


def write_readme_sheet(wb, obj):
    ws = wb.create_sheet("说明与数据来源")
    ws.column_dimensions["A"].width = 110
    lines = [
        ("《2026本科招生专业目录》数据说明", TITLE_FONT),
        ("", None),
        ("一、全国本科专业目录", Font(name="微软雅黑", bold=True, size=12)),
        ("· 共 883 个本科专业，覆盖哲学、经济学、法学、教育学、文学、历史学、理学、工学、"
         "农学、医学、管理学、艺术学、交叉学科 12 个学科门类。", CELL_FONT),
        ("· 权威依据：教育部《普通高等学校本科专业目录（2025年）》（教高函〔2025〕3号），"
         "含2023/2024/2025年新增专业，2026年招生执行。", CELL_FONT),
        ("· 官方原文：" + obj["source_authoritative_url"], CELL_FONT),
        ("· 专业代码后缀：T=特设专业，K=国家控制布点专业，TK=两者兼具。", CELL_FONT),
        ("", None),
        ("二、江苏省2026招生计划", Font(name="微软雅黑", bold=True, size=12)),
        ("· 江苏省教育考试院已于2026-06-23公布2026年普通高校招生计划（汇总数据见对应工作表）。", CELL_FONT),
        ("· 重要说明：江苏“逐所高校 × 逐个专业 × 选科要求 × 招生名额”的明细目录，"
         "仅刊印在《江苏招生考试2026招生计划专刊》并通过省考试院志愿填报系统查询，"
         "目前没有公开、可批量下载的结构化数据源。", Font(name="微软雅黑", size=10, color="C00000")),
        ("· 因此本表的江苏部分仅收录官方公布的权威汇总数字，未编造逐校逐专业明细，"
         "以免误导考生与家长。逐校明细请以《招生计划专刊》/省考试院系统为准。", CELL_FONT),
        ("· 来源：https://gaokao.chsi.com.cn/gkxx/zc/ss/202606/20260623/2293843281.html", CELL_FONT),
        ("", None),
        ("三、口径提醒", Font(name="微软雅黑", bold=True, size=12)),
        ("· 教育部专业目录是“全国本科专业总名录”，并非某一所高校或某一省的招生专业清单；"
         "各高校实际开设专业及其在各省的志愿填报代码、选科要求、计划数因校因省而异。", CELL_FONT),
        ("· 填报志愿请以考生所在省份当年发布的招生计划为准。", CELL_FONT),
    ]
    r = 1
    for text, font in lines:
        ws.cell(row=r, column=1, value=text)
        if font:
            ws.cell(row=r, column=1).font = font
        ws.cell(row=r, column=1).alignment = LEFT
        r += 1


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "本科招生专业目录_2026.xlsx"
    obj = json.loads(MAJORS_FILE.read_text(encoding="utf-8"))
    majors = obj["majors"]
    wb = Workbook()
    write_catalog_sheet(wb, majors)
    write_category_sheet(wb, majors)
    write_jiangsu_sheet(wb)
    write_readme_sheet(wb, obj)
    wb.save(out)
    print(f"已生成 Excel：{out}（专业 {len(majors)} 个，工作表 {len(wb.sheetnames)} 个）")
    print("工作表：", " / ".join(wb.sheetnames))


if __name__ == "__main__":
    main()
