# -*- coding: utf-8 -*-
"""根据 /tmp/zhuankan_parsed.json 生成《江苏2026招生计划·院校专业组》Excel。"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data" / "jiangsu_zhuankan_parsed_2026.json"
OUT = ROOT / "江苏2026招生计划_院校专业组.xlsx"

HF = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
TFONT = Font(name="微软雅黑", bold=True, size=14, color="1F4E78")
CFONT = Font(name="微软雅黑", size=10)
RED = Font(name="微软雅黑", size=10, color="C00000")
THIN = Side(style="thin", color="BFBFBF")
BD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEF = Alignment(horizontal="left", vertical="center", wrap_text=True)

CAT_ORDER = {"提前录取本科":0,"本科批":1,"艺术类":2,"体育类":3,"专科":4,"":9}

def head(ws,n):
    for c in range(1,n+1):
        cell=ws.cell(row=1,column=c); cell.fill=HF; cell.font=HFONT; cell.alignment=CEN; cell.border=BD

def detail(wb,rows):
    ws=wb.active; ws.title="江苏2026院校专业组明细"
    H=["序号","科类","类别","院校代号","院校名称","专业组代号","组号","选科要求","专业组(原文)",
       "专业代号","专业名称","计划数","学制","学费","数据来源页"]
    ws.append(H); head(ws,len(H))
    rows=sorted(rows,key=lambda r:(r.get("km",""),CAT_ORDER.get(r.get("cat",""),9),
                                   str(r.get("school_code") or ""),str(r.get("grp_code") or ""),
                                   str(r.get("sp_code") or "")))
    for i,r in enumerate(rows,1):
        grpno=(r["grp_code"][4:6] if r.get("grp_code") and len(r["grp_code"])>=6 else "")
        sel=r.get("grp_select") or ""
        km=r.get("km","")
        xk=(f"首选{km}，再选{sel}" if sel else (f"首选{km}" if km else ""))
        ws.append([i,km,r.get("cat",""),r.get("school_code",""),r.get("school",""),
                   r.get("grp_code",""),grpno,xk,r.get("grp_name",""),
                   r.get("sp_code",""),r.get("sp_name",""),r.get("plan"),r.get("len",""),
                   r.get("fee"),r.get("page","")])
    leftcols={5,9,11}
    for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=len(H)):
        for cell in row:
            cell.font=CFONT; cell.alignment=LEF if cell.column in leftcols else CEN
    for i,w in enumerate([6,6,12,9,22,11,6,18,28,8,28,7,6,8,9],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(H))}{ws.max_row}"
    return len(rows)

def groups_sheet(wb,groups):
    ws=wb.create_sheet("专业组汇总")
    H=["序号","科类","类别","院校代号","院校名称","专业组代号","组号","选科·备注",
       "组内专业数","组计划(表头)","专业计划合计","是否一致"]
    ws.append(H); head(ws,len(H))
    from collections import Counter
    spc=Counter((g['km'],g['cat'],g['school_code'],g['grp_code']) for g in groups)  # not used
    gs=sorted(groups,key=lambda g:(g.get("km",""),CAT_ORDER.get(g.get("cat",""),9),
                                   str(g.get("school_code") or ""),str(g.get("grp_code") or "")))
    for i,g in enumerate(gs,1):
        grpno=(g["grp_code"][4:6] if g.get("grp_code") and len(g["grp_code"])>=6 else "")
        m=g.get("match")
        ws.append([i,g.get("km",""),g.get("cat",""),g.get("school_code",""),g.get("school",""),
                   g.get("grp_code",""),grpno,g.get("grp_select",""),
                   g.get("n_major",""),g.get("grp_plan"),g.get("sp_plan_sum"),
                   "一致" if m else ("不一致" if m is False else "无表头")])
    for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=len(H)):
        for cell in row:
            cell.font=CFONT; cell.border=BD
            cell.alignment=LEF if cell.column in {5,8} else CEN
            if cell.column==12 and cell.value=="不一致": cell.font=RED
    for i,w in enumerate([6,6,12,9,24,11,6,22,10,11,11,9],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(H))}{ws.max_row}"

def readme(wb,meta):
    ws=wb.create_sheet("说明与数据来源"); ws.column_dimensions["A"].width=118
    L=[("《江苏2026招生计划·院校专业组》数据说明",TFONT),("",None),
       ("一、来源",Font(name='微软雅黑',bold=True,size=12)),
       ("· 来自江苏省教育考试院《江苏招生考试 2026 招生计划专刊》PDF（扫描图）经 OCR(RapidOCR/PP-OCR) 自动识别并结构化还原。",CFONT),
       (f"· 共识别数据页 {meta['data_pages']} 页，重建：专业行 {meta['n_rows']} 条，院校专业组 {meta['n_groups']} 个。",CFONT),
       ("· 层级：院校(代号+名称) → 院校专业组(代号+选科要求) → 组内专业(代号+名称+计划数+学制+学费)。",CFONT),
       ("",None),
       ("二、字段",Font(name='微软雅黑',bold=True,size=12)),
       ("· 院校代号/专业组代号/专业代号 均为江苏志愿填报官方代号；专业组代号=院校代号(4位)+组号(2位)。",CFONT),
       ("· 类别：提前录取本科 / 本科批 / 艺术类 / 体育类（专科一般在专刊下册，本册以本科为主）。",CFONT),
       ("",None),
       ("三、准确性与校验（重要）",Font(name='微软雅黑',bold=True,size=12)),
       (f"· 本表由扫描件 OCR 自动生成，可能存在个别识别误差（尤其代号/数字）。已做校验：对比每个专业组的"
        f"“表头计划数”与“组内专业计划之和”，不一致的在『专业组汇总』表标记为“不一致”，共 {meta['n_group_mismatch']} 个，请优先抽查这些组。",RED),
       ("· 正式填报务必以《招生计划专刊》纸质版 / 江苏省教育考试院志愿系统(gk.jseea.cn)逐条核对，本表仅供整理与检索参考。",RED),
      ]
    r=1
    for t,f in L:
        ws.cell(row=r,column=1,value=t)
        if f: ws.cell(row=r,column=1).font=f
        ws.cell(row=r,column=1).alignment=LEF; r+=1

def main():
    d=json.loads(SRC.read_text())
    # n_major per group
    from collections import Counter
    cnt=Counter((r['km'],r['cat'],r['school_code'],r['grp_code']) for r in d["rows"])
    for g in d["groups"]:
        g["n_major"]=cnt.get((g['km'],g['cat'],g['school_code'],g['grp_code']),0)
    wb=Workbook()
    n=detail(wb,d["rows"]); groups_sheet(wb,d["groups"])
    readme(wb,{"data_pages":d["data_pages"],"n_rows":d["n_rows"],
               "n_groups":d["n_groups"],"n_group_mismatch":d["n_group_mismatch"]})
    wb.save(OUT)
    print("saved",OUT,"明细",n,"组",len(d["groups"]))

if __name__=="__main__":
    main()
